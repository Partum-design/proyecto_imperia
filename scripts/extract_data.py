import json
from datetime import datetime, date
from collections import Counter, defaultdict
import openpyxl
import docx

wb = openpyxl.load_workbook('Base de datos Partum.xlsx', data_only=True)
ws_conf = wb['Conf_Listas']
ws_tasks = wb['Cronograma Mayo 2026']


def clean(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, str):
        t = v.strip()
        return t if t else None
    return v

employees = []
clients_from_conf = set()
for row in ws_conf.iter_rows(min_row=2, values_only=True):
    emp, email, client, dept = map(clean, row[:4])
    if client:
        clients_from_conf.add(client)
    if emp:
        employees.append({
            'name': emp,
            'email': email,
            'department': dept,
            'assignedClient': client,
        })

headers = [clean(c) for c in next(ws_tasks.iter_rows(min_row=1, max_row=1, values_only=True))]
raw_tasks = []
for row in ws_tasks.iter_rows(min_row=2, values_only=True):
    rec = {headers[i]: clean(row[i]) if i < len(row) else None for i in range(len(headers))}
    if not any(rec.values()):
        continue
    activity = rec.get('Actividad')
    client = rec.get('Cliente')
    collaborator = rec.get('Colaborador')
    if not (activity or client or collaborator):
        continue

    start_date = rec.get('Fecha de inicio')
    due_date = rec.get('Fecha de entrega')
    progress = rec.get('%Avance') or 0
    try:
        progress = float(progress)
    except Exception:
        progress = 0

    raw_tasks.append({
        'startDate': start_date,
        'dueDate': due_date,
        'month': rec.get('Mes'),
        'owner': collaborator,
        'email': rec.get('Correo'),
        'client': client,
        'department': rec.get('Departamento'),
        'activity': activity,
        'progress': max(0, min(progress, 1)),
        'status': rec.get('Estado '),
        'delay': rec.get('Retrasos '),
        'notes': rec.get('Notas'),
    })

# normalize and enrich task records
now = datetime(2026, 5, 19)
enriched_tasks = []
for i, t in enumerate(raw_tasks, start=1):
    due = None
    if t['dueDate']:
        try:
            due = datetime.fromisoformat(t['dueDate'][:19])
        except Exception:
            due = None

    days_to_due = None
    alert = 'Sin fecha'
    priority = 'media'
    if due:
        days_to_due = (due.date() - now.date()).days
        if days_to_due < 0:
            alert = f'Vencida hace {abs(days_to_due)} día(s)'
            priority = 'alta'
        elif days_to_due == 0:
            alert = 'Vence hoy'
            priority = 'alta'
        elif days_to_due <= 2:
            alert = f'Vence en {days_to_due} día(s)'
            priority = 'alta'
        elif days_to_due <= 7:
            alert = f'Vence en {days_to_due} día(s)'
            priority = 'media'
        else:
            alert = f'Tiempo suficiente ({days_to_due} día(s))'
            priority = 'baja'

    status = (t['status'] or 'Sin iniciar').strip()
    status_l = status.lower()
    if 'complet' in status_l or t['progress'] >= 1:
        bucket = 'Completado'
    elif 'progreso' in status_l:
        bucket = 'En progreso'
    elif 'iniciar' in status_l:
        bucket = 'Sin iniciar'
    else:
        bucket = status or 'Sin iniciar'

    enriched_tasks.append({
        'id': f'TSK-{i:04d}',
        **t,
        'statusBucket': bucket,
        'daysToDue': days_to_due,
        'alert': alert,
        'priority': priority,
    })

clients = sorted({c for c in clients_from_conf if c} | {t['client'] for t in enriched_tasks if t.get('client')})

# KPIs
status_counts = Counter(t['statusBucket'] for t in enriched_tasks)
priority_counts = Counter(t['priority'] for t in enriched_tasks)
department_counts = Counter(t['department'] or 'Sin departamento' for t in enriched_tasks)

owner_stats = defaultdict(lambda: {'tasks': 0, 'avgProgress': 0.0, 'completed': 0, 'overdue': 0})
for t in enriched_tasks:
    owner = t['owner'] or 'Sin asignar'
    s = owner_stats[owner]
    s['tasks'] += 1
    s['avgProgress'] += t['progress']
    if t['statusBucket'] == 'Completado':
        s['completed'] += 1
    if t['daysToDue'] is not None and t['daysToDue'] < 0 and t['statusBucket'] != 'Completado':
        s['overdue'] += 1

owner_stats_list = []
for name, s in owner_stats.items():
    owner_stats_list.append({
        'name': name,
        'tasks': s['tasks'],
        'avgProgress': round((s['avgProgress'] / s['tasks']) if s['tasks'] else 0, 4),
        'completed': s['completed'],
        'overdue': s['overdue'],
    })
owner_stats_list.sort(key=lambda x: x['tasks'], reverse=True)

# extract reminder script summary from docx
script_doc = docx.Document('JAVA.docx')
script_lines = [p.text.strip() for p in script_doc.paragraphs if p.text.strip()]

data = {
    'meta': {
        'projectName': 'Proyecto Imperia',
        'sourceWorkbook': 'Base de datos Partum.xlsx',
        'sourceDoc': 'JAVA.docx',
        'generatedAt': datetime.now().isoformat(timespec='seconds'),
        'records': {
            'employees': len(employees),
            'clients': len(clients),
            'tasks': len(enriched_tasks),
        }
    },
    'employees': employees,
    'clients': clients,
    'tasks': enriched_tasks,
    'kpis': {
        'status': status_counts,
        'priority': priority_counts,
        'departments': department_counts,
    },
    'ownerStats': owner_stats_list,
    'automationLogic': {
        'title': 'Recordatorios y calendario',
        'rules': [
            'Crear evento de calendario por tarea no completada',
            'Enviar correo si vence hoy, ya venció o vence en <=2 días',
            'Personalizar asunto y cuerpo con cliente, tarea, estado y fecha'
        ],
        'docSnippet': script_lines[:18],
    }
}

with open('data/data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f"Exported {len(enriched_tasks)} tasks, {len(clients)} clients, {len(employees)} employees to data/data.json")
